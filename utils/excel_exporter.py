"""
–≠–∫—Å–ø–æ—Ä—Ç–µ—Ä —Å—Ç–∞—Ç–∏—Å—Ç–∏–∫–∏ –≤ Excel —Ñ–∞–π–ª
"""
import asyncio
import logging
import os
from datetime import datetime
from typing import Dict, List

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

logger = logging.getLogger(__name__)


class ExcelExporter:
    """–ö–ª–∞—Å—Å –¥–ª—è —ç–∫—Å–ø–æ—Ä—Ç–∞ —Å—Ç–∞—Ç–∏—Å—Ç–∏–∫–∏ –≤ Excel —Ñ–∞–π–ª"""

    # –¶–≤–µ—Ç–∞ –¥–ª—è —Ñ–æ—Ä–º–∞—Ç–∏—Ä–æ–≤–∞–Ω–∏—è (hex)
    COLORS = {
        'header': '4285F4',    # –°–∏–Ω–∏–π Google
        'subheader': 'F0F0F0', # –°–≤–µ—Ç–ª–æ-—Å–µ—Ä—ã–π
        'warning': 'FFF4CC',   # –°–≤–µ—Ç–ª–æ-–∂–µ–ª—Ç—ã–π
        'error': 'FFE6E6',     # –°–≤–µ—Ç–ª–æ-–∫—Ä–∞—Å–Ω—ã–π
        'success': 'D9EAD3',   # –°–≤–µ—Ç–ª–æ-–∑–µ–ª–µ–Ω—ã–π
    }

    def __init__(self):
        """–ò–Ω–∏—Ü–∏–∞–ª–∏–∑–∞—Ü–∏—è"""
        # –°–æ–∑–¥–∞–µ–º –ø–∞–ø–∫—É –¥–ª—è —ç–∫—Å–ø–æ—Ä—Ç–∞, –µ—Å–ª–∏ –Ω–µ—Ç
        self.base_dir = os.path.join(os.getcwd(), 'exports')
        os.makedirs(self.base_dir, exist_ok=True)

    async def export_stats(self, stats_data: Dict) -> str:
        """
        –°–æ–∑–¥–∞–µ—Ç Excel —Ñ–∞–π–ª (–∞—Å–∏–Ω—Ö—Ä–æ–Ω–Ω–∞—è –æ–±–µ—Ä—Ç–∫–∞)
        """
        loop = asyncio.get_running_loop()
        # –ó–∞–ø—É—Å–∫–∞–µ–º —Ç—è–∂–µ–ª—É—é –∑–∞–¥–∞—á—É –≤ –æ—Ç–¥–µ–ª—å–Ω–æ–º –ø–æ—Ç–æ–∫–µ
        return await loop.run_in_executor(None, self._create_excel_file, stats_data)

    def _create_excel_file(self, stats_data: Dict) -> str:
        """–°–∏–Ω—Ö—Ä–æ–Ω–Ω–∞—è —Ñ—É–Ω–∫—Ü–∏—è —Å–æ–∑–¥–∞–Ω–∏—è —Ñ–∞–π–ª–∞"""
        try:
            wb = Workbook()
            ws = wb.active
            ws.title = "–°—Ç–∞—Ç–∏—Å—Ç–∏–∫–∞"

            # –í–ê–ñ–ù–û: –≠—Ç–∏ —Ñ—É–Ω–∫—Ü–∏–∏ —Ç–µ–ø–µ—Ä—å —Å–∏–Ω—Ö—Ä–æ–Ω–Ω—ã–µ (–±–µ–∑ async)
            self._fill_worksheet(ws, stats_data)
            self._apply_formatting(ws)
            self._auto_resize_columns(ws)

            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')

            # –£–±–∏—Ä–∞–µ–º –æ–ø–∞—Å–Ω—ã–µ —Å–∏–º–≤–æ–ª—ã –∏–∑ –∏–º–µ–Ω–∏ —Ñ–∞–π–ª–∞
            safe_title = "".join([c for c in stats_data['channel'].title if c.isalnum() or c in (' ', '_')]).strip()
            channel_name = safe_title.replace(' ', '_')

            filename = f"stats_{channel_name}_{timestamp}.xlsx"
            filepath = os.path.join(self.base_dir, filename)

            wb.save(filepath)
            logger.info(f"Stats exported to Excel: {filepath}")
            return filepath

        except Exception as e:
            logger.error(f"Error creating Excel file: {e}", exc_info=True)
            raise

    def _fill_worksheet(self, ws, stats_data: Dict):
        """–ó–∞–ø–æ–ª–Ω—è–µ—Ç –ª–∏—Å—Ç –¥–∞–Ω–Ω—ã–º–∏ (Synchronous)"""
        channel = stats_data['channel']
        timestamp = stats_data['timestamp']
        events = stats_data['events']

        row = 1

        # === –ó–ê–ì–û–õ–û–í–û–ö ===
        ws.cell(row, 1, f"üìä –°—Ç–∞—Ç–∏—Å—Ç–∏–∫–∞: {channel.title}")
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)

        row += 1
        ws.cell(row, 1, f"–ù–∞ –º–æ–º–µ–Ω—Ç: {timestamp.strftime('%d.%m.%Y %H:%M')}")
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)

        row += 2

        # –ó–∞–≥–æ–ª–æ–≤–∫–∏ –∫–æ–ª–æ–Ω–æ–∫
        headers = ["–°–æ–±—ã—Ç–∏–µ", "–¢–∏–ø", "–î–µ–¥–ª–∞–π–Ω", "–°—Ç–∞—Ç—É—Å", "–ú–∞–≥–∞–∑–∏–Ω/–ü–æ–ª—å–∑–æ–≤–∞—Ç–µ–ª—å", "–î–µ—Ç–∞–ª–∏"]
        for col, header in enumerate(headers, start=1):
            ws.cell(row, col, header)

        row += 1
        start_data_row = row  # –ó–∞–ø–æ–º–∏–Ω–∞–µ–º –Ω–∞—á–∞–ª–æ –¥–∞–Ω–Ω—ã—Ö

        # === –û–ë–´–ß–ù–´–ï –°–û–ë–´–¢–ò–Ø ===
        for item in events.get('regular', []):
            event = item['event']
            not_submitted = item['not_submitted']

            for store_id, users_list in not_submitted:
                ws.cell(row, 1, event.keyword)
                ws.cell(row, 2, "–û–±—ã—á–Ω–æ–µ")
                ws.cell(row, 3, event.deadline_time.strftime('%H:%M'))
                ws.cell(row, 4, "‚ùå –ù–µ —Å–¥–∞–ª–∏")
                ws.cell(row, 5, self._format_store_for_excel(store_id, users_list))
                ws.cell(row, 6, f"–¢—Ä–µ–±—É–µ—Ç—Å—è: {event.min_photos} —Ñ–æ—Ç–æ")
                row += 1

        # === –í–†–ï–ú–ï–ù–ù–´–ï –°–û–ë–´–¢–ò–Ø ===
        for item in events.get('temp', []):
            temp_event = item['event']
            not_submitted = item['not_submitted']

            for store_id, users_list in not_submitted:
                ws.cell(row, 1, temp_event.keyword)
                ws.cell(row, 2, "–í—Ä–µ–º–µ–Ω–Ω–æ–µ")
                ws.cell(row, 3, temp_event.deadline_time.strftime('%H:%M'))
                ws.cell(row, 4, "‚ùå –ù–µ —Å–¥–∞–ª–∏")
                ws.cell(row, 5, self._format_store_for_excel(store_id, users_list))
                ws.cell(row, 6, "–£–¥–∞–ª–∏—Ç—Å—è –≤ 23:59")
                row += 1

        # === CHECKOUT –°–û–ë–´–¢–ò–Ø ===
        for item in events.get('checkout', []):
            cev = item['event']
            checkout_stats = item['stats']

            # –ù–µ —Å–¥–∞–ª–∏ –ø–µ—Ä–≤—ã–π —ç—Ç–∞–ø
            for store_id, users_list in checkout_stats.get('not_submitted_first', []):
                ws.cell(row, 1, cev.first_keyword)
                ws.cell(row, 2, "Checkout (1 —ç—Ç–∞–ø)")
                ws.cell(row, 3, cev.first_deadline_time.strftime('%H:%M'))
                ws.cell(row, 4, "‚ö†Ô∏è –ù–µ —Å–¥–∞–ª–∏ 1 —ç—Ç–∞–ø")
                ws.cell(row, 5, self._format_store_for_excel(store_id, users_list))
                ws.cell(row, 6, "–ù—É–∂–µ–Ω —Å–ø–∏—Å–æ–∫ –∫–∞—Ç–µ–≥–æ—Ä–∏–π")
                row += 1

            # –ù–µ –Ω–∞—á–∞–ª–∏ –≤—Ç–æ—Ä–æ–π —ç—Ç–∞–ø
            for store_id, users_list in checkout_stats.get('not_submitted_second', []):
                ws.cell(row, 1, cev.second_keyword)
                ws.cell(row, 2, "Checkout (2 —ç—Ç–∞–ø)")
                ws.cell(row, 3, cev.second_deadline_time.strftime('%H:%M'))
                ws.cell(row, 4, "‚ö†Ô∏è –ù–µ –Ω–∞—á–∞–ª–∏ 2 —ç—Ç–∞–ø")
                ws.cell(row, 5, self._format_store_for_excel(store_id, users_list))
                ws.cell(row, 6, "1 —ç—Ç–∞–ø —Å–¥–∞–Ω, –∂–¥–µ–º —Ñ–æ—Ç–æ")
                row += 1

            # –°–¥–∞–ª–∏ —á–∞—Å—Ç–∏—á–Ω–æ
            for store_id, users_list, remaining in checkout_stats.get('partial_second', []):
                ws.cell(row, 1, cev.second_keyword)
                ws.cell(row, 2, "Checkout (2 —ç—Ç–∞–ø)")
                ws.cell(row, 3, cev.second_deadline_time.strftime('%H:%M'))
                ws.cell(row, 4, "‚ö†Ô∏è –ß–∞—Å—Ç–∏—á–Ω–æ —Å–¥–∞–Ω–æ")
                ws.cell(row, 5, self._format_store_for_excel(store_id, users_list))
                ws.cell(row, 6, f"–û—Å—Ç–∞–ª–æ—Å—å: {', '.join(remaining)}")
                row += 1

            # –ù–µ —Å–¥–∞–ª–∏ –Ω–∏—á–µ–≥–æ
            for store_id, users_list in checkout_stats.get('not_submitted_anything', []):
                ws.cell(row, 1, cev.first_keyword)
                ws.cell(row, 2, "Checkout")
                ws.cell(row, 3, cev.first_deadline_time.strftime('%H:%M'))
                ws.cell(row, 4, "‚ùå –ù–µ —Å–¥–∞–ª–∏ –Ω–∏—á–µ–≥–æ")
                ws.cell(row, 5, self._format_store_for_excel(store_id, users_list))
                ws.cell(row, 6, "–î–µ–¥–ª–∞–π–Ω 1 —ç—Ç–∞–ø–∞ –ø—Ä–æ—à–µ–ª")
                row += 1

        # === NOTEXT –°–û–ë–´–¢–ò–Ø ===
        for item in events.get('notext', []):
            notext_event = item['event']
            not_submitted = item['not_submitted']

            for store_id, users_list in not_submitted:
                ws.cell(row, 1, "–§–æ—Ç–æ –±–µ–∑ —Ç–µ–∫—Å—Ç–∞")
                ws.cell(row, 2, "NoText")
                time_range = f"{notext_event.deadline_start.strftime('%H:%M')}-{notext_event.deadline_end.strftime('%H:%M')}"
                ws.cell(row, 3, time_range)
                ws.cell(row, 4, "‚ùå –ù–µ —Å–¥–∞–ª–∏")
                ws.cell(row, 5, self._format_store_for_excel(store_id, users_list))
                ws.cell(row, 6, "–ù—É–∂–Ω–æ —Ñ–æ—Ç–æ –≤ —É–∫–∞–∑–∞–Ω–Ω—ã–π –ø–µ—Ä–∏–æ–¥")
                row += 1

        # === KEYWORD –°–û–ë–´–¢–ò–Ø ===
        for item in events.get('keyword', []):
            keyword_event = item['event']
            not_submitted = item['not_submitted']

            for store_id, users_list in not_submitted:
                ws.cell(row, 1, keyword_event.keyword)
                ws.cell(row, 2, "Keyword")
                time_range = f"{keyword_event.deadline_start.strftime('%H:%M')}-{keyword_event.deadline_end.strftime('%H:%M')}"
                ws.cell(row, 3, time_range)
                ws.cell(row, 4, "‚ùå –ù–µ —Å–¥–∞–ª–∏")
                ws.cell(row, 5, self._format_store_for_excel(store_id, users_list))
                ws.cell(row, 6, f"–ù—É–∂–Ω–æ —Å–æ–æ–±—â–µ–Ω–∏–µ —Å '{keyword_event.keyword}'")
                row += 1

        # –ï—Å–ª–∏ –Ω–µ—Ç –¥–∞–Ω–Ω—ã—Ö
        if row == start_data_row:
            ws.cell(row, 1, "üéâ –í—Å–µ –æ—Ç—á–µ—Ç—ã —Å–¥–∞–Ω—ã!")
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)

        # –°–æ—Ö—Ä–∞–Ω—è–µ–º –ø–æ—Å–ª–µ–¥–Ω—é—é —Å—Ç—Ä–æ–∫—É –¥–ª—è —Ñ–æ—Ä–º–∞—Ç–∏—Ä–æ–≤–∞–Ω–∏—è
        ws._last_data_row = row

    def _apply_formatting(self, ws):
        """–ü—Ä–∏–º–µ–Ω—è–µ—Ç —Ñ–æ—Ä–º–∞—Ç–∏—Ä–æ–≤–∞–Ω–∏–µ –∫ —Ç–∞–±–ª–∏—Ü–µ (Synchronous)"""

        # –®—Ä–∏—Ñ—Ç—ã
        header_font = Font(name='Arial', size=14, bold=True, color='FFFFFF')
        subheader_font = Font(name='Arial', size=11, bold=True)
        normal_font = Font(name='Arial', size=10)

        # –ó–∞–ª–∏–≤–∫–∏
        header_fill = PatternFill(start_color=self.COLORS['header'], end_color=self.COLORS['header'], fill_type='solid')
        subheader_fill = PatternFill(start_color=self.COLORS['subheader'], end_color=self.COLORS['subheader'], fill_type='solid')
        warning_fill = PatternFill(start_color=self.COLORS['warning'], end_color=self.COLORS['warning'], fill_type='solid')
        error_fill = PatternFill(start_color=self.COLORS['error'], end_color=self.COLORS['error'], fill_type='solid')
        success_fill = PatternFill(start_color=self.COLORS['success'], end_color=self.COLORS['success'], fill_type='solid')

        # –í—ã—Ä–∞–≤–Ω–∏–≤–∞–Ω–∏–µ
        center_alignment = Alignment(horizontal='center', vertical='center')
        left_alignment = Alignment(horizontal='left', vertical='center')
        wrap_alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)

        # –ì—Ä–∞–Ω–∏—Ü—ã
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        # –ó–∞–≥–æ–ª–æ–≤–æ–∫ (—Å—Ç—Ä–æ–∫–∞ 1 –∏ 2)
        # –ü—Ä–∏–º–µ–Ω—è–µ–º —Å—Ç–∏–ª—å –∫ –ø–µ—Ä–≤–æ–π —è—á–µ–π–∫–µ –æ–±—ä–µ–¥–∏–Ω–µ–Ω–Ω–æ–≥–æ –¥–∏–∞–ø–∞–∑–æ–Ω–∞
        cell_1 = ws.cell(1, 1)
        cell_1.font = header_font
        cell_1.fill = header_fill
        cell_1.alignment = left_alignment

        cell_2 = ws.cell(2, 1)
        cell_2.font = Font(name='Arial', size=12, bold=False, color='FFFFFF')
        cell_2.fill = header_fill
        cell_2.alignment = left_alignment

        # –ó–∞–≥–æ–ª–æ–≤–∫–∏ –∫–æ–ª–æ–Ω–æ–∫ (—Å—Ç—Ä–æ–∫–∞ 4)
        for col in range(1, 7):
            cell = ws.cell(4, col)
            cell.font = subheader_font
            cell.fill = subheader_fill
            cell.alignment = center_alignment
            cell.border = thin_border

        # –î–∞–Ω–Ω—ã–µ (—Å 5 —Å—Ç—Ä–æ–∫–∏)
        last_row = getattr(ws, '_last_data_row', ws.max_row)

        for row_idx in range(5, last_row + 1):

            # –û–ø—Ä–µ–¥–µ–ª—è–µ–º —Ü–≤–µ—Ç —Å—Ç—Ä–æ–∫–∏ –ø–æ —Å—Ç–∞—Ç—É—Å—É
            status_cell = ws.cell(row_idx, 4)
            status_text = status_cell.value or ""

            row_fill = None
            if "‚ùå" in status_text:
                row_fill = error_fill
            elif "‚ö†Ô∏è" in status_text:
                row_fill = warning_fill
            elif "‚úÖ" in status_text or "üéâ" in status_text:
                row_fill = success_fill

            for col_idx in range(1, 7):
                cell = ws.cell(row_idx, col_idx)
                cell.font = normal_font
                cell.border = thin_border

                # –í—ã—Ä–∞–≤–Ω–∏–≤–∞–Ω–∏–µ
                if col_idx in [1, 2, 5, 6]:  # –¢–µ–∫—Å—Ç–æ–≤—ã–µ –∫–æ–ª–æ–Ω–∫–∏
                    cell.alignment = left_alignment if col_idx != 2 else center_alignment
                else:  # –î–µ–¥–ª–∞–π–Ω, —Å—Ç–∞—Ç—É—Å
                    cell.alignment = center_alignment

                # –ü–µ—Ä–µ–Ω–æ—Å —Ç–µ–∫—Å—Ç–∞ –¥–ª—è –¥–ª–∏–Ω–Ω—ã—Ö –∫–æ–ª–æ–Ω–æ–∫
                if col_idx in [5, 6]:
                    cell.alignment = wrap_alignment

                # –ü—Ä–∏–º–µ–Ω—è–µ–º –∑–∞–ª–∏–≤–∫—É –µ—Å–ª–∏ –æ–Ω–∞ –æ–ø—Ä–µ–¥–µ–ª–µ–Ω–∞
                if row_fill:
                    cell.fill = row_fill

    def _auto_resize_columns(self, ws):
        """–ê–≤—Ç–æ–º–∞—Ç–∏—á–µ—Å–∫–∏–π –ø–æ–¥–±–æ—Ä —à–∏—Ä–∏–Ω—ã –∫–æ–ª–æ–Ω–æ–∫"""
        column_widths = {
            'A': 25,  # –°–æ–±—ã—Ç–∏–µ
            'B': 20,  # –¢–∏–ø
            'C': 15,  # –î–µ–¥–ª–∞–π–Ω
            'D': 20,  # –°—Ç–∞—Ç—É—Å
            'E': 30,  # –ú–∞–≥–∞–∑–∏–Ω/–ü–æ–ª—å–∑–æ–≤–∞—Ç–µ–ª—å
            'F': 35,  # –î–µ—Ç–∞–ª–∏
        }

        for col_letter, width in column_widths.items():
            ws.column_dimensions[col_letter].width = width

    def _format_store_for_excel(self, store_id: str, users_list: List) -> str:
        """–§–æ—Ä–º–∞—Ç–∏—Ä—É–µ—Ç —É–ø–æ–º–∏–Ω–∞–Ω–∏–µ –º–∞–≥–∞–∑–∏–Ω–∞ –¥–ª—è Excel"""
        if store_id.startswith("no_store_"):
            user = users_list[0]
            if user.username:
                return f"@{user.username}"
            return user.full_name or f"ID:{user.telegram_id}"

        # –î–ª—è –º–∞–≥–∞–∑–∏–Ω–∞: "MSK-001 (@user1, @user2)"
        usernames = []
        for user in users_list:
            if user.username:
                usernames.append(f"@{user.username}")
            else:
                usernames.append(user.full_name or f"ID:{user.telegram_id}")

        if usernames:
            return f"{store_id} ({', '.join(usernames)})"
        return store_id