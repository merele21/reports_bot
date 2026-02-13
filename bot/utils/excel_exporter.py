"""
Экспортер статистики в Excel файл
"""
import asyncio
import os
import logging
from typing import Dict, List
from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

logger = logging.getLogger(__name__)


class ExcelExporter:
    """Класс для экспорта статистики в Excel файл"""

    # Цвета для форматирования (hex)
    COLORS = {
        'header': '4285F4',  # Синий Google
        'subheader': 'F0F0F0',  # Светло-серый
        'warning': 'FFF4CC',  # Светло-желтый
        'error': 'FFE6E6',  # Светло-красный
        'success': 'D9EAD3',  # Светло-зеленый
    }

    def __init__(self):
        """Инициализация"""
        # Создаем папку для экспорта, если нет
        self.base_dir = os.path.join(os.getcwd(), 'exports')
        os.makedirs(self.base_dir, exist_ok=True)

    async def export_stats(self, stats_data: Dict) -> str:
        """
        Создает Excel файл (асинхронная обертка)
        """
        loop = asyncio.get_running_loop()
        # Запускаем тяжелую задачу в отдельном потоке
        return await loop.run_in_executor(None, self._create_excel_file, stats_data)

    def _create_excel_file(self, stats_data: Dict) -> str:
        """Синхронная функция создания файла"""
        try:
            wb = Workbook()
            ws = wb.active
            ws.title = "Статистика"

            self._fill_worksheet(ws, stats_data)
            self._apply_formatting(ws, stats_data)
            self._auto_resize_columns(ws)

            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            # Убираем опасные символы из имени файла
            safe_title = "".join([c for c in stats_data['channel'].title if c.isalnum() or c in (' ', '_')]).strip()
            channel_name = safe_title.replace(' ', '_')

            filename = f"stats_{channel_name}_{timestamp}.xlsx"
            filepath = os.path.join(self.base_dir, filename)

            wb.save(filepath)
            logger.info(f"Stats exported to Excel: {filepath}")
            return filepath

        except Exception as e:
            logger.error(f"Error creating Excel file: {e}", exc_info=True)
            raise

    async def _fill_worksheet(self, ws, stats_data: Dict):
        """Заполняет лист данными"""
        channel = stats_data['channel']
        timestamp = stats_data['timestamp']
        events = stats_data['events']

        row = 1

        # === ЗАГОЛОВОК ===
        ws.cell(row, 1, f"📊 Статистика: {channel.title}")
        ws.cell(row, 2, f"На момент: {timestamp.strftime('%d.%m.%Y %H:%M')}")
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=1)
        ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=6)
        row += 2

        # Заголовки колонок
        headers = ["Событие", "Тип", "Дедлайн", "Статус", "Магазин/Пользователь", "Детали"]
        for col, header in enumerate(headers, start=1):
            ws.cell(row, col, header)
        row += 1

        start_data_row = row  # Запоминаем начало данных

        # === ОБЫЧНЫЕ СОБЫТИЯ ===
        for item in events['regular']:
            event = item['event']
            not_submitted = item['not_submitted']

            for store_id, users_list in not_submitted:
                ws.cell(row, 1, event.keyword)
                ws.cell(row, 2, "Обычное")
                ws.cell(row, 3, event.deadline_time.strftime('%H:%M'))
                ws.cell(row, 4, "❌ Не сдали")
                ws.cell(row, 5, self._format_store_for_excel(store_id, users_list))
                ws.cell(row, 6, f"Требуется: {event.min_photos} фото")
                row += 1

        # === ВРЕМЕННЫЕ СОБЫТИЯ ===
        for item in events['temp']:
            temp_event = item['event']
            not_submitted = item['not_submitted']

            for store_id, users_list in not_submitted:
                ws.cell(row, 1, temp_event.keyword)
                ws.cell(row, 2, "Временное")
                ws.cell(row, 3, temp_event.deadline_time.strftime('%H:%M'))
                ws.cell(row, 4, "❌ Не сдали")
                ws.cell(row, 5, self._format_store_for_excel(store_id, users_list))
                ws.cell(row, 6, "Удалится в 23:59")
                row += 1

        # === CHECKOUT СОБЫТИЯ ===
        for item in events['checkout']:
            cev = item['event']
            checkout_stats = item['stats']

            # Не сдали первый этап
            for store_id, users_list in checkout_stats['not_submitted_first']:
                ws.cell(row, 1, cev.first_keyword)
                ws.cell(row, 2, "Checkout (1 этап)")
                ws.cell(row, 3, cev.first_deadline_time.strftime('%H:%M'))
                ws.cell(row, 4, "⚠️ Не сдали 1 этап")
                ws.cell(row, 5, self._format_store_for_excel(store_id, users_list))
                ws.cell(row, 6, "Нужен список категорий")
                row += 1

            # Не начали второй этап
            for store_id, users_list in checkout_stats['not_submitted_second']:
                ws.cell(row, 1, cev.second_keyword)
                ws.cell(row, 2, "Checkout (2 этап)")
                ws.cell(row, 3, cev.second_deadline_time.strftime('%H:%M'))
                ws.cell(row, 4, "⚠️ Не начали 2 этап")
                ws.cell(row, 5, self._format_store_for_excel(store_id, users_list))
                ws.cell(row, 6, "1 этап сдан, ждем фото")
                row += 1

            # Сдали частично
            for store_id, users_list, remaining in checkout_stats['partial_second']:
                ws.cell(row, 1, cev.second_keyword)
                ws.cell(row, 2, "Checkout (2 этап)")
                ws.cell(row, 3, cev.second_deadline_time.strftime('%H:%M'))
                ws.cell(row, 4, "⚠️ Частично сдано")
                ws.cell(row, 5, self._format_store_for_excel(store_id, users_list))
                ws.cell(row, 6, f"Осталось: {', '.join(remaining)}")
                row += 1

            # Не сдали ничего
            for store_id, users_list in checkout_stats['not_submitted_anything']:
                ws.cell(row, 1, cev.first_keyword)
                ws.cell(row, 2, "Checkout")
                ws.cell(row, 3, cev.first_deadline_time.strftime('%H:%M'))
                ws.cell(row, 4, "❌ Не сдали ничего")
                ws.cell(row, 5, self._format_store_for_excel(store_id, users_list))
                ws.cell(row, 6, "Дедлайн 1 этапа прошел")
                row += 1

        # === NOTEXT СОБЫТИЯ ===
        for item in events['notext']:
            notext_event = item['event']
            not_submitted = item['not_submitted']

            for store_id, users_list in not_submitted:
                ws.cell(row, 1, "Фото без текста")
                ws.cell(row, 2, "NoText")
                time_range = f"{notext_event.deadline_start.strftime('%H:%M')}-{notext_event.deadline_end.strftime('%H:%M')}"
                ws.cell(row, 3, time_range)
                ws.cell(row, 4, "❌ Не сдали")
                ws.cell(row, 5, self._format_store_for_excel(store_id, users_list))
                ws.cell(row, 6, "Нужно фото в указанный период")
                row += 1

        # === KEYWORD СОБЫТИЯ ===
        for item in events['keyword']:
            keyword_event = item['event']
            not_submitted = item['not_submitted']

            for store_id, users_list in not_submitted:
                ws.cell(row, 1, keyword_event.keyword)
                ws.cell(row, 2, "Keyword")
                time_range = f"{keyword_event.deadline_start.strftime('%H:%M')}-{keyword_event.deadline_end.strftime('%H:%M')}"
                ws.cell(row, 3, time_range)
                ws.cell(row, 4, "❌ Не сдали")
                ws.cell(row, 5, self._format_store_for_excel(store_id, users_list))
                ws.cell(row, 6, f"Нужно сообщение с '{keyword_event.keyword}'")
                row += 1

        # Если нет данных
        if row == start_data_row:
            ws.cell(row, 1, "🎉 Все отчеты сданы!")
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)

        # Сохраняем последнюю строку для форматирования
        ws._last_data_row = row

    async def _apply_formatting(self, ws, stats_data: Dict):
        """Применяет форматирование к таблице"""

        # Шрифты
        header_font = Font(name='Arial', size=14, bold=True, color='FFFFFF')
        subheader_font = Font(name='Arial', size=11, bold=True)
        normal_font = Font(name='Arial', size=10)

        # Заливки
        header_fill = PatternFill(start_color=self.COLORS['header'], end_color=self.COLORS['header'], fill_type='solid')
        subheader_fill = PatternFill(start_color=self.COLORS['subheader'], end_color=self.COLORS['subheader'], fill_type='solid')
        warning_fill = PatternFill(start_color=self.COLORS['warning'], end_color=self.COLORS['warning'], fill_type='solid')
        error_fill = PatternFill(start_color=self.COLORS['error'], end_color=self.COLORS['error'], fill_type='solid')
        success_fill = PatternFill(start_color=self.COLORS['success'], end_color=self.COLORS['success'], fill_type='solid')

        # Выравнивание
        center_alignment = Alignment(horizontal='center', vertical='center')
        left_alignment = Alignment(horizontal='left', vertical='center')
        wrap_alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)

        # Границы
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        # Заголовок (строка 1)
        for col in range(1, 7):
            cell = ws.cell(1, col)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = left_alignment

        # Заголовки колонок (строка 3)
        for col in range(1, 7):
            cell = ws.cell(3, col)
            cell.font = subheader_font
            cell.fill = subheader_fill
            cell.alignment = center_alignment
            cell.border = thin_border

        # Данные (с 4 строки)
        last_row = getattr(ws, '_last_data_row', ws.max_row)

        for row_idx in range(4, last_row + 1):
            for col_idx in range(1, 7):
                cell = ws.cell(row_idx, col_idx)
                cell.font = normal_font
                cell.border = thin_border

                # Выравнивание
                if col_idx in [1, 2, 5, 6]:  # Текстовые колонки
                    cell.alignment = left_alignment if col_idx != 2 else center_alignment
                else:  # Дедлайн, статус
                    cell.alignment = center_alignment

                # Перенос текста для длинных колонок
                if col_idx in [5, 6]:
                    cell.alignment = wrap_alignment

            # Цветовое кодирование по статусу
            status_cell = ws.cell(row_idx, 4)
            status_text = status_cell.value or ""

            if "❌" in status_text:
                fill = error_fill
            elif "⚠️" in status_text:
                fill = warning_fill
            elif "✅" in status_text or "🎉" in status_text:
                fill = success_fill
            else:
                continue

            # Применяем заливку ко всей строке
            for col_idx in range(1, 7):
                ws.cell(row_idx, col_idx).fill = fill

    def _auto_resize_columns(self, ws):
        """Автоматический подбор ширины колонок"""
        column_widths = {
            'A': 25,  # Событие
            'B': 20,  # Тип
            'C': 12,  # Дедлайн
            'D': 20,  # Статус
            'E': 30,  # Магазин/Пользователь
            'F': 35,  # Детали
        }

        for col_letter, width in column_widths.items():
            ws.column_dimensions[col_letter].width = width

    def _format_store_for_excel(self, store_id: str, users_list: List) -> str:
        """Форматирует упоминание магазина для Excel"""
        if store_id.startswith("no_store_"):
            user = users_list[0]
            if user.username:
                return f"@{user.username}"
            return user.full_name or f"ID:{user.telegram_id}"

        # Для магазина: "MSK-001 (@user1, @user2)"
        usernames = []
        for user in users_list:
            if user.username:
                usernames.append(f"@{user.username}")
            else:
                usernames.append(user.full_name or f"ID:{user.telegram_id}")

        if usernames:
            return f"{store_id} ({', '.join(usernames)})"
        return store_id